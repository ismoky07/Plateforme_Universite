"""
Générateur de fichiers Excel pour l'administration
"""

import pandas as pd
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule

from admin_config import VALIDATION_STATUS, ANOMALY_TYPES
from admin_utils import get_candidature_details


def export_all_candidatures_excel(candidatures):
    """Exporte toutes les candidatures vers Excel"""
    
    # Créer le workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Feuille 1: Vue d'ensemble
    create_overview_sheet(wb, candidatures)
    
    # Feuille 2: Détails des candidatures
    create_details_sheet(wb, candidatures)
    
    # Feuille 3: Statistiques
    create_statistics_sheet(wb, candidatures)
    
    # Sauvegarder en mémoire
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def create_overview_sheet(wb, candidatures):
    """Crée la feuille de vue d'ensemble"""
    ws = wb.create_sheet("Vue d'ensemble", 0)
    
    # Préparer les données
    data = []
    for candidature in candidatures:
        details = get_candidature_details(candidature)
        
        status_info = VALIDATION_STATUS.get(details['status'], VALIDATION_STATUS['en_attente'])
        
        data.append({
            'Candidat': f"{details['candidat']['prenom']} {details['candidat']['nom']}",
            'Email': details['candidat']['email'],
            'Niveau': details['candidat']['niveau_etude'],
            'Notes saisies': details['statistiques'].get('nombre_notes', 0),
            'Moyenne': details['statistiques'].get('moyenne_generale', 0),
            'Documents': details['statistiques'].get('nombre_documents', 0),
            'Date candidature': details['soumission'].get('date', '')[:10] if details['soumission'].get('date') else '',
            'Statut': status_info['name'],
            'Validateur': details.get('validator', ''),
            'Date validation': details.get('validation_date', '')[:10] if details.get('validation_date') else '',
            'Référence': details['soumission'].get('reference', '')
        })
    
    # Convertir en DataFrame
    df = pd.DataFrame(data)
    
    # Ajouter au worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Mise en forme
    apply_header_formatting(ws)
    apply_status_formatting(ws, df)
    auto_adjust_columns(ws)


def create_details_sheet(wb, candidatures):
    """Crée la feuille des détails candidatures"""
    ws = wb.create_sheet("Détails candidatures")
    
    # En-têtes
    headers = [
        'Candidat', 'Email', 'Téléphone', 'Niveau', 'Référence',
        'Date candidature', 'Statut', 'Validateur', 'Date validation', 'Commentaires',
        'Matière', 'Note', 'Coefficient', 'Période', 'Année'
    ]
    
    ws.append(headers)
    
    # Données détaillées
    for candidature in candidatures:
        details = get_candidature_details(candidature)
        
        # Informations de base
        base_info = [
            f"{details['candidat']['prenom']} {details['candidat']['nom']}",
            details['candidat']['email'],
            details['candidat'].get('telephone', ''),
            details['candidat']['niveau_etude'],
            details['soumission'].get('reference', ''),
            details['soumission'].get('date', '')[:10] if details['soumission'].get('date') else '',
            VALIDATION_STATUS.get(details['status'], VALIDATION_STATUS['en_attente'])['name'],
            details.get('validator', ''),
            details.get('validation_date', '')[:10] if details.get('validation_date') else '',
            details.get('comments', '')
        ]
        
        # Si pas de notes, ajouter une ligne avec les infos de base
        if not details['notes']:
            ws.append(base_info + ['', '', '', '', ''])
        else:
            # Une ligne par note
            for note in details['notes']:
                row = base_info + [
                    note.get('matiere', ''),
                    note.get('note', ''),
                    note.get('coefficient', ''),
                    note.get('periode', ''),
                    note.get('annee', '')
                ]
                ws.append(row)
    
    # Mise en forme
    apply_header_formatting(ws)
    auto_adjust_columns(ws)


def create_statistics_sheet(wb, candidatures):
    """Crée la feuille des statistiques"""
    ws = wb.create_sheet("Statistiques")
    
    # Statistiques générales
    total_candidatures = len(candidatures)
    
    # Compter par statut
    status_counts = {}
    for candidature in candidatures:
        details = get_candidature_details(candidature)
        status = details['status']
        status_counts[status] = status_counts.get(status, 0) + 1
    
    # Compter par niveau
    niveau_counts = {}
    for candidature in candidatures:
        details = get_candidature_details(candidature)
        niveau = details['candidat']['niveau_etude']
        niveau_counts[niveau] = niveau_counts.get(niveau, 0) + 1
    
    # Écrire les statistiques
    ws.append(['STATISTIQUES GÉNÉRALES'])
    ws.append([])
    ws.append(['Total candidatures', total_candidatures])
    ws.append([])
    
    ws.append(['RÉPARTITION PAR STATUT'])
    for status, count in status_counts.items():
        status_name = VALIDATION_STATUS.get(status, {'name': status})['name']
        ws.append([status_name, count, f"{count/total_candidatures*100:.1f}%"])
    
    ws.append([])
    ws.append(['RÉPARTITION PAR NIVEAU'])
    for niveau, count in niveau_counts.items():
        ws.append([niveau, count, f"{count/total_candidatures*100:.1f}%"])
    
    # Moyennes par niveau
    ws.append([])
    ws.append(['MOYENNES PAR NIVEAU'])
    ws.append(['Niveau', 'Moyenne générale', 'Nombre candidats'])
    
    for niveau in niveau_counts.keys():
        candidatures_niveau = [
            c for c in candidatures 
            if get_candidature_details(c)['candidat']['niveau_etude'] == niveau
        ]
        
        moyennes = []
        for candidature in candidatures_niveau:
            details = get_candidature_details(candidature)
            moyenne = details['statistiques'].get('moyenne_generale', 0)
            if moyenne > 0:
                moyennes.append(moyenne)
        
        moyenne_niveau = sum(moyennes) / len(moyennes) if moyennes else 0
        
        ws.append([niveau, f"{moyenne_niveau:.2f}", len(candidatures_niveau)])
    
    # Mise en forme
    apply_statistics_formatting(ws)
    auto_adjust_columns(ws)


def export_candidature_excel(candidature):
    """Exporte une candidature individuelle vers Excel"""
    details = get_candidature_details(candidature)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Feuille informations candidat
    ws_info = wb.create_sheet("Informations candidat", 0)
    
    # Informations personnelles
    ws_info.append(['INFORMATIONS PERSONNELLES'])
    ws_info.append([])
    ws_info.append(['Nom', details['candidat']['nom']])
    ws_info.append(['Prénom', details['candidat']['prenom']])
    ws_info.append(['Email', details['candidat']['email']])
    ws_info.append(['Téléphone', details['candidat'].get('telephone', 'Non renseigné')])
    ws_info.append(['Niveau d\'étude', details['candidat']['niveau_etude']])
    ws_info.append([])
    
    # Informations de candidature
    ws_info.append(['CANDIDATURE'])
    ws_info.append([])
    ws_info.append(['Référence', details['soumission'].get('reference', '')])
    ws_info.append(['Date de soumission', details['soumission'].get('date', '')])
    ws_info.append(['Statut', VALIDATION_STATUS.get(details['status'], VALIDATION_STATUS['en_attente'])['name']])
    ws_info.append(['Validateur', details.get('validator', '')])
    ws_info.append(['Date de validation', details.get('validation_date', '')])
    ws_info.append(['Commentaires', details.get('comments', '')])
    
    # Feuille des notes
    if details['notes']:
        ws_notes = wb.create_sheet("Notes saisies")
        
        # En-têtes
        ws_notes.append(['Matière', 'Note (/20)', 'Coefficient', 'Période', 'Année'])
        
        # Données
        for note in details['notes']:
            ws_notes.append([
                note.get('matiere', ''),
                note.get('note', ''),
                note.get('coefficient', ''),
                note.get('periode', ''),
                note.get('annee', '')
            ])
        
        # Calculs
        ws_notes.append([])
        ws_notes.append(['STATISTIQUES'])
        ws_notes.append(['Nombre de notes', len(details['notes'])])
        ws_notes.append(['Moyenne générale', details['statistiques'].get('moyenne_generale', 0)])
        
        apply_header_formatting(ws_notes)
        auto_adjust_columns(ws_notes)
    
    # Mise en forme
    apply_info_formatting(ws_info)
    auto_adjust_columns(ws_info)
    
    # Sauvegarder
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def export_complete_analysis_excel(candidature, comparison_result):
    """Exporte l'analyse complète avec comparaison OCR"""
    details = get_candidature_details(candidature)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Feuille 1: Résumé de l'analyse
    create_analysis_summary_sheet(wb, details, comparison_result)
    
    # Feuille 2: Comparaison détaillée
    create_comparison_sheet(wb, comparison_result)
    
    # Feuille 3: Anomalies détectées
    create_anomalies_sheet(wb, comparison_result)
    
    # Feuille 4: Notes originales
    create_original_notes_sheet(wb, details)
    
    # Sauvegarder
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()


def create_analysis_summary_sheet(wb, details, comparison_result):
    """Crée la feuille de résumé d'analyse"""
    ws = wb.create_sheet("Résumé analyse", 0)
    
    # Informations candidat
    ws.append(['CANDIDAT'])
    ws.append([])
    ws.append(['Nom', f"{details['candidat']['prenom']} {details['candidat']['nom']}"])
    ws.append(['Email', details['candidat']['email']])
    ws.append(['Niveau', details['candidat']['niveau_etude']])
    ws.append(['Référence', details['soumission'].get('reference', '')])
    ws.append([])
    
    # Statistiques de comparaison
    stats = comparison_result['statistiques']
    
    ws.append(['RÉSULTATS DE LA COMPARAISON'])
    ws.append([])
    ws.append(['Notes manuelles', stats['total_notes_manuelles']])
    ws.append(['Correspondances trouvées', stats['correspondances_trouvees']])
    ws.append(['Taux de correspondance', f"{stats['taux_correspondance']:.1%}"])
    ws.append(['Anomalies détectées', stats['anomalies_detectees']])
    ws.append(['Confiance OCR moyenne', f"{stats['confiance_moyenne_ocr']:.1%}"])
    ws.append(['Tolérance utilisée', f"{stats['tolerance_utilisee']} points"])
    ws.append([])
    
    # Recommandation
    ws.append(['RECOMMANDATION'])
    ws.append([])
    
    if stats['anomalies_detectees'] == 0:
        recommendation = "✅ VALIDÉE - Aucune anomalie détectée"
        color = "00FF00"
    elif stats['anomalies_detectees'] <= 2 and stats['taux_correspondance'] > 0.8:
        recommendation = "⚠️ À VÉRIFIER - Anomalies mineures détectées"
        color = "FFA500"
    else:
        recommendation = "❌ À REJETER - Anomalies majeures détectées"
        color = "FF0000"
    
    ws.append(['Statut recommandé', recommendation])
    
    # Mise en forme
    apply_analysis_formatting(ws, color)
    auto_adjust_columns(ws)


def create_comparison_sheet(wb, comparison_result):
    """Crée la feuille de comparaison détaillée"""
    ws = wb.create_sheet("Comparaison détaillée")
    
    # En-têtes
    headers = [
        'Matière Saisie', 'Note Saisie', 'Coeff Saisi',
        'Matière OCR', 'Note OCR', 'Coeff OCR',
        'Écart Note', 'Confiance OCR', 'Similarité', 'Statut', 'Anomalie'
    ]
    ws.append(headers)
    
    # Données de comparaison
    for comp in comparison_result['comparisons']:
        row = [
            comp['matiere_manuelle'],
            comp['note_manuelle'],
            comp['coefficient_manuel'],
            comp.get('matiere_ocr', 'N/A'),
            comp.get('note_ocr', 'N/A'),
            comp.get('coefficient_ocr', 'N/A'),
            comp.get('ecart_note', 'N/A'),
            f"{comp.get('confidence_ocr', 0):.1%}" if comp.get('confidence_ocr') else 'N/A',
            f"{comp.get('similarite_matiere', 0):.1%}" if comp.get('similarite_matiere') else 'N/A',
            comp['statut'],
            'OUI' if comp.get('anomalie', False) else 'NON'
        ]
        ws.append(row)
    
    # Mise en forme
    apply_comparison_formatting(ws)
    auto_adjust_columns(ws)


def create_anomalies_sheet(wb, comparison_result):
    """Crée la feuille des anomalies"""
    ws = wb.create_sheet("Anomalies détectées")
    
    if not comparison_result['anomalies']:
        ws.append(['Aucune anomalie détectée'])
        return
    
    # En-têtes
    headers = ['Type', 'Matière', 'Note Saisie', 'Note OCR', 'Écart', 'Sévérité', 'Description']
    ws.append(headers)
    
    # Anomalies
    for anomaly in comparison_result['anomalies']:
        anomaly_info = ANOMALY_TYPES.get(anomaly['type'], {})
        
        row = [
            anomaly_info.get('name', anomaly['type']),
            anomaly['matiere'],
            anomaly.get('note_manuelle', 'N/A'),
            anomaly.get('note_ocr', 'N/A'),
            anomaly.get('ecart', 'N/A'),
            anomaly['severity'].upper(),
            anomaly_info.get('description', '')
        ]
        ws.append(row)
    
    # Mise en forme
    apply_anomalies_formatting(ws)
    auto_adjust_columns(ws)


def create_original_notes_sheet(wb, details):
    """Crée la feuille des notes originales"""
    ws = wb.create_sheet("Notes originales")
    
    # En-têtes
    ws.append(['Matière', 'Note (/20)', 'Coefficient', 'Période', 'Année'])
    
    # Notes
    for note in details['notes']:
        ws.append([
            note.get('matiere', ''),
            note.get('note', ''),
            note.get('coefficient', ''),
            note.get('periode', ''),
            note.get('annee', '')
        ])
    
    # Mise en forme
    apply_header_formatting(ws)
    auto_adjust_columns(ws)


# Fonctions de mise en forme
def apply_header_formatting(ws):
    """Applique la mise en forme aux en-têtes"""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def apply_status_formatting(ws, df):
    """Applique la mise en forme selon le statut"""
    status_colors = {
        'En attente': 'FFF3CD',
        'Validée': 'D4EDDA', 
        'Rejetée': 'F8D7DA',
        'En cours d\'examen': 'D1ECF1'
    }
    
    # Trouver la colonne statut
    status_col = None
    for idx, col in enumerate(df.columns, 1):
        if 'Statut' in col:
            status_col = idx
            break
    
    if status_col:
        for row_idx, status in enumerate(df['Statut'], 2):
            if status in status_colors:
                cell = ws.cell(row=row_idx, column=status_col)
                cell.fill = PatternFill(start_color=status_colors[status], 
                                       end_color=status_colors[status], 
                                       fill_type="solid")


def apply_info_formatting(ws):
    """Applique la mise en forme aux feuilles d'informations"""
    title_font = Font(bold=True, size=14)
    
    # Mise en forme des titres
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and cell.value.isupper():
            cell.font = title_font


def apply_analysis_formatting(ws, recommendation_color):
    """Applique la mise en forme à l'analyse"""
    title_font = Font(bold=True, size=14)
    recommendation_fill = PatternFill(start_color=recommendation_color, 
                                     end_color=recommendation_color, 
                                     fill_type="solid")
    
    # Titres
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and cell.value.isupper():
            cell.font = title_font
        
        # Recommandation
        if cell.value == 'Statut recommandé' and len(row) > 1:
            row[1].fill = recommendation_fill


def apply_comparison_formatting(ws):
    """Applique la mise en forme à la comparaison"""
    apply_header_formatting(ws)
    
    # Mise en forme conditionnelle pour les anomalies
    anomaly_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        if len(row) > 10 and row[10].value == 'OUI':  # Colonne Anomalie
            for cell in row:
                cell.fill = anomaly_fill


def apply_anomalies_formatting(ws):
    """Applique la mise en forme aux anomalies"""
    apply_header_formatting(ws)
    
    severity_colors = {
        'HIGH': 'F8D7DA',
        'MEDIUM': 'FFF3CD',
        'LOW': 'D1ECF1'
    }
    
    for row in ws.iter_rows(min_row=2):
        if len(row) > 5:
            severity = row[5].value
            if severity in severity_colors:
                for cell in row:
                    cell.fill = PatternFill(start_color=severity_colors[severity], 
                                           end_color=severity_colors[severity], 
                                           fill_type="solid")


def apply_statistics_formatting(ws):
    """Applique la mise en forme aux statistiques"""
    title_font = Font(bold=True, size=12)
    
    for row in ws.iter_rows():
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and (
            'STATISTIQUES' in cell.value or 'RÉPARTITION' in cell.value or 'MOYENNES' in cell.value
        ):
            cell.font = title_font


def auto_adjust_columns(ws):
    """Ajuste automatiquement la largeur des colonnes"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width